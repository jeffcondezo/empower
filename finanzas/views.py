from django.views.generic import DetailView, ListView, TemplateView, RedirectView
from django.shortcuts import redirect, HttpResponse
import json
from decimal import Decimal

# Mixin Import-->
from maestro.mixin import BasicEMixin
# Mixin Import<--

# Extra python features-->
from datetime import datetime
# Extra python features<--
import datetime as dt

# Model import-->
from finanzas.models import Jornada, DetalleJornada, CuentaCliente, CuentaProveedor, PagoCliente, PagoProveedor
from ventas.models import Venta
from compras.models import Compra
# Model import<--
from maestro.utils import empresa_list
# Forms import-->
from finanzas.forms import JornadaFiltroForm, DetalleJornadaCreateForm,\
    JornadaCreateForm, CuentaClienteFiltroForm, PagoClienteCreateForm, CuentaProveedorFiltroForm,\
    PagoProveedorCreateForm, PagoVentaForm, PagoCompraForm
# Forms import<--

from openpyxl import Workbook,load_workbook
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.styles import Border, Side


class JornadaListView(BasicEMixin, ListView):

    template_name = 'finanzas/jornada-list.html'
    model = Jornada
    nav_name = 'nav_jornada'
    view_name = 'jornada'
    action_name = 'leer'

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        context['jornada_filtro'] = JornadaFiltroForm(self.request.GET, user=self.request.user)
        context['jornada_open'] = JornadaCreateForm(user=self.request.user)
        return context

    def get_queryset(self):
        query = super().get_queryset()
        caja = self.request.GET.getlist('caja')
        alt = 0
        if len(caja) > 0:
            query = query.filter(caja__in=caja)
            alt += 1
        if 'fechahora_inicio1' in self.request.GET and 'fechahora_inicio2' in self.request.GET:
            if self.request.GET['fechahora_inicio1'] != '' and self.request.GET['fechahora_inicio2'] != '':
                fecha_inicio = datetime.strptime(self.request.GET['fechahora_inicio1'], '%d/%m/%Y %H:%M')
                fecha_fin = datetime.strptime(self.request.GET['fechahora_inicio2'], '%d/%m/%Y %H:%M')
                query = query.filter(fechahora_inicio__gte=fecha_inicio, fechahora_inicio__lte=fecha_fin)
                alt += 1
        if 'monto1' in self.request.GET or 'monto2' in self.request.GET:
            monto1 = self.request.GET['monto1']
            monto2 = self.request.GET['monto2']
            if monto1 == '':
                query = query.filter(monto_actual__lte=monto2)
                alt += 1
            elif monto2 == '':
                query = query.filter(monto_actual__gte=monto1)
                alt += 1
            else:
                query = query.filter(monto_actual__gte=monto1, monto_actual__lte=monto2)
                alt += 1
        if alt == 0:
            query = Jornada.objects.filter(estado=True)
        query.filter(caja__sucursal__empresa__in=empresa_list(self.request.user))
        return query


class JornadaDetailView(BasicEMixin, DetailView):

    template_name = 'finanzas/jornada-detail.html'
    model = Jornada
    nav_name = 'nav_jornada'
    view_name = 'jornada'
    action_name = 'leer'

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        context['detalle'] = DetalleJornada.objects.filter(jornada=self.kwargs['pk'])
        context['djornada_create'] = DetalleJornadaCreateForm()
        return context


class DetalleJornadaCreateView(RedirectView):

    url = '/finanzas/jornada/'
    view_name = 'jornada'
    action_name = 'crear'

    def get_redirect_url(self, *args, **kwargs):
        jornada = Jornada.objects.get(pk=self.kwargs['jornada'])
        form = DetalleJornadaCreateForm(self.request.POST)
        if form.is_valid():
            djornada = form.save(commit=False)
            djornada.asignado = self.request.user
            djornada.jornada = jornada
            djornada.save()
            if djornada.tipo == '1':
                jornada.monto_actual += djornada.monto
            else:
                jornada.monto_actual -= djornada.monto
            jornada.save()
        url = self.url + str(self.kwargs['jornada'])
        return url


class JornadaCloseView(RedirectView):

    url = '/finanzas/jornada/'
    view_name = 'jornada'
    action_name = 'cerrar'

    def get_redirect_url(self, *args, **kwargs):
        jornada = Jornada.objects.get(pk=self.kwargs['jornada'])
        jornada.estado = False
        jornada.save()
        url = self.url + str(self.kwargs['jornada'])
        return url


class JornadaCreateView(RedirectView):

    url = '/finanzas/jornada/'
    view_name = 'jornada'
    action_name = 'abrir'

    def get_redirect_url(self, *args, **kwargs):
        form = JornadaCreateForm(self.request.POST, user=self.request.user)
        if form.is_valid():
            try:
                jornada = Jornada.objects.get(caja=form.cleaned_data['caja'], estado=True)
            except Jornada.DoesNotExist:
                jornada = form.save(commit=False)
                jornada.asignado_inicio = self.request.user
                jornada.monto_actual = jornada.monto_apertura
                jornada.save()
        url = self.url + str(jornada.id)
        return url


class CuentaClienteListView(BasicEMixin, ListView):

    template_name = 'finanzas/cuentacliente-list.html'
    model = CuentaCliente
    nav_name = 'nav_cuentacliente'
    view_name = 'cuenta_cliente'
    action_name = 'leer'

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        context['cuentacliente_filtro'] = CuentaClienteFiltroForm(self.request.GET, user=self.request.user)
        return context

    def get_queryset(self):
        query = super().get_queryset()
        cliente = self.request.GET.getlist('cliente')
        duracion = self.request.GET.getlist('duracion')
        estado = self.request.GET.getlist('estado')
        tipo = self.request.GET.getlist('tipo')
        alt = 0
        if len(cliente) > 0:
            alt += 1
            query = query.filter(cliente__in=cliente)
        if len(duracion) > 0:
            alt += 1
            query = query.filter(duracion__in=duracion)
        if len(estado) > 0:
            alt += 1
            query = query.filter(estado__in=estado)
        if len(tipo) > 0:
            alt += 1
            query = query.filter(tipo__in=tipo)
        if 'fechahora_caducidad1' in self.request.GET and 'fechahora_caducidad2' in self.request.GET:
            if self.request.GET['fechahora_caducidad1'] != '' and self.request.GET['fechahora_caducidad2'] != '':
                alt += 1
                fecha_inicio = datetime.strptime(self.request.GET['fechahora_caducidad1'], '%d/%m/%Y %H:%M')
                fecha_fin = datetime.strptime(self.request.GET['fechahora_caducidad2'], '%d/%m/%Y %H:%M')
                query = query.filter(fechahora_caducidad__gte=fecha_inicio, fechahora_caducidad__lte=fecha_fin)
        if 'monto_amortizado1' in self.request.GET or 'monto_amortizado2' in self.request.GET:
            monto_amortizado1 = self.request.GET['monto_amortizado1']
            monto_amortizado2 = self.request.GET['monto_amortizado2']
            if monto_amortizado1 == '' and monto_amortizado2 != '':
                alt += 1
                query = query.filter(monto_amortizado__lte=monto_amortizado2)
            elif monto_amortizado2 == '' and monto_amortizado1 != '':
                alt += 1
                query = query.filter(monto_amortizado__gte=monto_amortizado1)
            elif monto_amortizado1 != '' and monto_amortizado2 != '':
                alt += 1
                query = query.filter(monto_amortizado__gte=monto_amortizado1, monto_amortizado__lte=monto_amortizado2)
        if 'monto_total1' in self.request.GET or 'monto_total2' in self.request.GET:
            monto_total1 = self.request.GET['monto_total1']
            monto_total2 = self.request.GET['monto_total2']
            if monto_total1 == '' and monto_total2 != '':
                alt += 1
                query = query.filter(monto_total__lte=monto_total2)
            elif monto_total2 == '' and monto_total1 != '':
                alt += 1
                query = query.filter(monto_total__gte=monto_total1)
            elif monto_total1 != '' and monto_total2 != '':
                alt += 1
                query = query.filter(monto_total__gte=monto_total1, monto_total__lte=monto_total2)
        if alt == 0:
            query = query.filter(estado='1')
        query = query.filter(cliente__empresa__in=empresa_list(self.request.user))
        return query


class CuentaClienteDetailView(BasicEMixin, DetailView):

    template_name = 'finanzas/cuentacliente-detail.html'
    model = CuentaCliente
    nav_name = 'nav_cuentacliente'
    view_name = 'cuenta_cliente'
    action_name = 'leer'

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        context['detalle'] = PagoCliente.objects.filter(cuentacliente=self.kwargs['pk'])
        context['pago_create'] = PagoClienteCreateForm()
        return context


class PagoClienteCreateView(RedirectView):

    url = '/finanzas/cuentacliente/'
    view_name = 'pago_cliente'
    action_name = 'crear'

    def get_redirect_url(self, *args, **kwargs):
        cuentacliente = CuentaCliente.objects.get(pk=self.kwargs['cuentacliente'])
        form = PagoClienteCreateForm(self.request.POST)
        if form.is_valid():
            pago = form.save(commit=False)
            pago.asignado = self.request.user
            pago.cuentacliente = cuentacliente
            pago.save()
            cuentacliente.monto_amortizado += pago.monto
            cuentacliente.monto_deuda -= pago.monto
            if cuentacliente.monto_deuda == 0:
                cuentacliente.estado = '2'
            cuentacliente.save()
        url = self.url + str(self.kwargs['cuentacliente'])
        return url


class CuentaProveedorListView(BasicEMixin, ListView):

    template_name = 'finanzas/cuentaproveedor-list.html'
    model = CuentaProveedor
    nav_name = 'nav_cuentaproveedor'
    view_name = 'cuenta_proveedor'
    action_name = 'leer'

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        context['cuentaproveedor_filtro'] = CuentaProveedorFiltroForm(self.request.GET, user=self.request.user)
        return context

    def get_queryset(self):
        query = super().get_queryset()
        cliente = self.request.GET.getlist('cliente')
        duracion = self.request.GET.getlist('duracion')
        estado = self.request.GET.getlist('estado')
        tipo = self.request.GET.getlist('tipo')
        alt = 0
        if len(cliente) > 0:
            alt += 1
            query = query.filter(cliente__in=cliente)
        if len(duracion) > 0:
            alt += 1
            query = query.filter(duracion__in=duracion)
        if len(estado) > 0:
            alt += 1
            query = query.filter(estado__in=estado)
        if len(tipo) > 0:
            alt += 1
            query = query.filter(tipo__in=tipo)
        if 'fechahora_caducidad1' in self.request.GET and 'fechahora_caducidad2' in self.request.GET:
            if self.request.GET['fechahora_caducidad1'] != '' and self.request.GET['fechahora_caducidad2'] != '':
                alt += 1
                fecha_inicio = datetime.strptime(self.request.GET['fechahora_caducidad1'], '%d/%m/%Y %H:%M')
                fecha_fin = datetime.strptime(self.request.GET['fechahora_caducidad2'], '%d/%m/%Y %H:%M')
                query = query.filter(fechahora_caducidad__gte=fecha_inicio, fechahora_caducidad__lte=fecha_fin)
        if 'monto_amortizado1' in self.request.GET or 'monto_amortizado2' in self.request.GET:
            monto_amortizado1 = self.request.GET['monto_amortizado1']
            monto_amortizado2 = self.request.GET['monto_amortizado2']
            if monto_amortizado1 == '' and monto_amortizado2 != '':
                alt += 1
                query = query.filter(monto_amortizado__lte=monto_amortizado2)
            elif monto_amortizado2 == '' and monto_amortizado1 != '':
                alt += 1
                query = query.filter(monto_amortizado__gte=monto_amortizado1)
            elif monto_amortizado1 != '' and monto_amortizado2 != '':
                alt += 1
                query = query.filter(monto_amortizado__gte=monto_amortizado1, monto_amortizado__lte=monto_amortizado2)
        if 'monto_total1' in self.request.GET or 'monto_total2' in self.request.GET:
            monto_total1 = self.request.GET['monto_total1']
            monto_total2 = self.request.GET['monto_total2']
            if monto_total1 == '':
                alt += 1
                query = query.filter(monto_total__lte=monto_total2)
            elif monto_total2 == '':
                alt += 1
                query = query.filter(monto_total__gte=monto_total1)
            else:
                alt += 1
                query = query.filter(monto_total__gte=monto_total1, monto_total__lte=monto_total2)
        if alt == 0:
            query = query.filter(estado='1')
        return query


class CuentaProveedorDetailView(BasicEMixin, DetailView):

    template_name = 'finanzas/cuentaproveedor-detail.html'
    model = CuentaProveedor
    nav_name = 'nav_cuentaproveedor'
    view_name = 'cuenta_proveedor'
    action_name = 'leer'

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        context['detalle'] = PagoProveedor.objects.filter(cuentaproveedor=self.kwargs['pk'])
        context['pago_create'] = PagoProveedorCreateForm()
        return context


class PagoProveedorCreateView(RedirectView):

    url = '/finanzas/cuentaproveedor/'
    view_name = 'pago_proveedor'
    action_name = 'crear'

    def get_redirect_url(self, *args, **kwargs):
        cuentaproveedor = CuentaProveedor.objects.get(pk=self.kwargs['cuentaproveedor'])
        form = PagoProveedorCreateForm(self.request.POST)
        if form.is_valid():
            pago = form.save(commit=False)
            pago.asignado = self.request.user
            pago.cuentaproveedor = cuentaproveedor
            pago.save()
            cuentaproveedor.monto_amortizado += pago.monto
            cuentaproveedor.monto_deuda -= pago.monto
            if cuentaproveedor.monto_deuda == 0:
                cuentaproveedor.estado = '2'
            cuentaproveedor.save()
        url = self.url + str(self.kwargs['cuentaproveedor'])
        return url


class VentaPagoView(RedirectView):

    url = '/ventas/venta/'
    view_name = 'finanzas'
    action_name = 'venta_pago'

    def get_redirect_url(self, *args, **kwargs):
        venta = Venta.objects.get(pk=self.kwargs['venta'])
        form = PagoVentaForm(self.request.POST, instance=venta)
        if not venta.is_pagado:
            if form.is_valid():
                if venta.tipo_pago == '2':
                    if venta.cliente is None:
                        url = self.url + str(venta.id) + '/?incidencias=' + json.dumps(
                            [['3', 'No se puede pagar al credito sin cliente']])
                        return url
                    elif venta.total_final > venta.cliente.credito_disponible:
                        url = self.url + str(venta.id) + '/?incidencias=' + json.dumps(
                            [['3', 'El cliente no tiene suficiente linea de credito disponible.']])
                        return url
                try:
                    jornada = Jornada.objects.get(caja=form.cleaned_data['caja'], estado=True)
                except Jornada.DoesNotExist:
                    url = self.url + str(venta.id) + '/?incidencias='+json.dumps(
                        [['3', 'La caja está cerrada, no se pudo concretar el pago.']])
                    return url
                venta = form.save(commit=False)
                venta.is_pagado = True
                fecha_actual = datetime.now()
                if form.cleaned_data['duracion'] == '1':
                    fecha_final = fecha_actual + dt.timedelta(days=7)
                elif form.cleaned_data['duracion'] == '2':
                    fecha_final = fecha_actual + dt.timedelta(days=14)
                elif form.cleaned_data['duracion'] == '3':
                    fecha_final = fecha_actual + dt.timedelta(days=21)
                elif form.cleaned_data['duracion'] == '4':
                    fecha_final = fecha_actual + dt.timedelta(days=30)
                tipo = form.cleaned_data['tipo_pago']
                pago = form.cleaned_data['pago']
                if venta.cliente is not None:
                    if tipo == '2':
                        if venta.total_final == pago:
                            venta.tipo_pago = '1'
                            venta.estado_pago = '2'
                            estado = '2'
                        elif venta.total_final > pago:
                            venta.estado_pago = '1'
                            estado = '1'
                    else:
                        venta.estado_pago = '2'
                        estado = '2'
                    cuentacliente = CuentaCliente(duracion=form.cleaned_data['duracion'], tipo=tipo, venta=venta,
                                                  estado=estado, fechahora_caducidad=fecha_final,
                                                  monto_total=venta.total_con_descuento, monto_amortizado=pago,
                                                  monto_deuda=float(venta.total_con_descuento) - pago,
                                                  cliente=venta.cliente)
                    cuentacliente.save()
                    PagoCliente(tipo='1', monto=pago, cuentacliente=cuentacliente, asignado=self.request.user,
                                recibo=form.cleaned_data['recibo'], venta=venta).save()
                else:
                    venta.tipo_pago = '1'
                    venta.estado_pago = '2'
                jornada.monto_actual += Decimal(pago)
                jornada.save()
                DetalleJornada(jornada=jornada, tipo='1', target=venta.id, monto=pago, descripcion='Pago Venta',
                               asignado=self.request.user).save()
                venta.save()
            else:
                return HttpResponse(form.errors)
            url = self.url + str(venta.id) + '/edit'
        else:
            url = '/ventas/venta'
        return url


class CompraPagoView(RedirectView):

    url = '/compras/compra/'
    view_name = 'finanzas'
    action_name = 'compra_pago'

    def get_redirect_url(self, *args, **kwargs):
        compra = Compra.objects.get(pk=self.kwargs['compra'])
        form = PagoCompraForm(self.request.POST, instance=compra)
        if not compra.is_financiado:
            if form.is_valid():
                try:
                    jornada = Jornada.objects.get(caja=form.cleaned_data['caja'], estado=True)
                except Jornada.DoesNotExist:
                    url = self.url + str(compra.id) + '/?incidencias='+json.dumps([
                        ['3', 'La caja está cerrada, no se pudo concretar el pago.']])
                    return url
                pago = form.cleaned_data['pago']
                if form.cleaned_data['tipo_pago'] == '1':
                    if jornada.monto_actual < pago:
                        url = self.url + str(compra.id) + '/?incidencias='+json.dumps([
                            ['3', 'Saldo insuficiente en Caja']])
                        return url
                    descripcion = 'Pago Compra desde Caja'
                    jornada.monto_actual -= Decimal(pago)
                    jornada.save()
                else:
                    descripcion = 'Pago Compra Libre'
                compra = form.save(commit=False)
                compra.is_financiado = True
                fecha_actual = datetime.now()
                fecha_final = fecha_actual
                if form.cleaned_data['duracion'] == '1':
                    fecha_final = fecha_actual + dt.timedelta(days=7)
                elif form.cleaned_data['duracion'] == '2':
                    fecha_final = fecha_actual + dt.timedelta(days=14)
                elif form.cleaned_data['duracion'] == '3':
                    fecha_final = fecha_actual + dt.timedelta(days=21)
                elif form.cleaned_data['duracion'] == '4':
                    fecha_final = fecha_actual + dt.timedelta(days=30)
                tipo = form.cleaned_data['tipo_pago']
                estado = ''
                if tipo == '2':
                    if compra.total_final == pago:
                        compra.tipo_pago = '1'
                        compra.estado_pago = '2'
                        estado = '2'
                    elif compra.total_final > pago:
                        compra.estado_pago = '1'
                        estado = '1'
                else:
                    compra.estado_pago = '2'
                    estado = '2'
                cuentaproveedor = CuentaProveedor(duracion=form.cleaned_data['duracion'], tipo=tipo, compra=compra,
                                                  estado=estado, fechahora_caducidad=fecha_final,
                                                  monto_total=compra.total_inc_flete, monto_amortizado=pago,
                                                  monto_deuda=float(compra.total_inc_flete) - pago,
                                                  proveedor=compra.proveedor)
                cuentaproveedor.save()
                PagoProveedor(tipo='1', monto=pago, cuentaproveedor=cuentaproveedor, asignado=self.request.user,
                              recibo=form.cleaned_data['recibo'], compra=compra).save()
                DetalleJornada(jornada=jornada, tipo='2', target=compra.id, monto=pago, descripcion=descripcion,
                               asignado=self.request.user).save()
                compra.save()
            else:
                return HttpResponse(form.errors)
            url = self.url + str(compra.id)
        else:
            url = '/compras/compra'
        return url


def reporte_jornada(request,id):
    if request.POST['tipo_movimiento'] == '3':
        libro = Workbook()
        libro = load_workbook("./finanza_detalle.xlsx")
        h = libro.get_sheet_by_name("Hoja1")
        fecha_inicio = datetime.strptime(request.POST['fechahora_inicio3'], '%d/%m/%Y %H:%M')
        fecha_fin = datetime.strptime(request.POST['fechahora_inicio4'], '%d/%m/%Y %H:%M')

        detalle=DetalleJornada.objects.filter(fechahora__gte=fecha_inicio, fechahora__lte=fecha_fin)
        i=6
        a=0
        total=0
        for d in detalle:
            i = i + 1
            a=a+1

            h.cell(row=i, column=2).value = a
            h.cell(row=i, column=3).value = d.fechahora
            h.cell(row=i, column=4).value = d.get_tipo_display()
            h.cell(row=i, column=5).value = d.target
            h.cell(row=i, column=6).value = d.monto
            h.cell(row=i, column=7).value = d.descripcion
            total = total + d.monto

            ### BORDER CADA FILA ###
            h.cell(row=i, column=2).border = Border(top=Side(border_style='thin', color='FF000000'),
                right=Side(border_style='thin', color='FF000000'),
                bottom=Side(border_style='thin', color='FF000000'),
                left=Side(border_style='thin', color='FF000000'))

            h.cell(row=i, column=3).border = Border(top=Side(border_style='thin', color='FF000000'),
                right=Side(border_style='thin', color='FF000000'),
                bottom=Side(border_style='thin', color='FF000000'),
                left=Side(border_style='thin', color='FF000000'))

            h.cell(row=i, column=4).border = Border(top=Side(border_style='thin', color='FF000000'),
                right=Side(border_style='thin', color='FF000000'),
                bottom=Side(border_style='thin', color='FF000000'),
                left=Side(border_style='thin', color='FF000000'))

            h.cell(row=i, column=5).border = Border(top=Side(border_style='thin', color='FF000000'),
                right=Side(border_style='thin', color='FF000000'),
                bottom=Side(border_style='thin', color='FF000000'),
                left=Side(border_style='thin', color='FF000000'))

            h.cell(row=i, column=6).border = Border(top=Side(border_style='thin', color='FF000000'),
                right=Side(border_style='thin', color='FF000000'),
                bottom=Side(border_style='thin', color='FF000000'),
                left=Side(border_style='thin', color='FF000000'))

            h.cell(row=i, column=7).border = Border(top=Side(border_style='thin', color='FF000000'),
                right=Side(border_style='thin', color='FF000000'),
                bottom=Side(border_style='thin', color='FF000000'),
                left=Side(border_style='thin', color='FF000000'))


        i = i + 2
        h.cell(row=i, column=5).value = "TOTAL"
        h.cell(row=i, column=6).value = total


        response = HttpResponse(content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename=detallejornada.xls'
        libro.save(response)
        return response



    elif request.POST['tipo_movimiento'] == '4':

        libro = Workbook()
        libro = load_workbook("./ingreso_vs_egreso.xlsx")
        h = libro.get_sheet_by_name("Hoja1")

        fecha_inicio = datetime.strptime(request.POST['fechahora_inicio1'], '%d/%m/%Y %H:%M')
        fecha_fin = datetime.strptime(request.POST['fechahora_inicio2'], '%d/%m/%Y %H:%M')

        detalle_ingreso=DetalleJornada.objects.filter(tipo=1,fechahora__gte=fecha_inicio, fechahora__lte=fecha_fin)
        detalle_egreso=DetalleJornada.objects.filter(tipo=2,fechahora__gte=fecha_inicio, fechahora__lte=fecha_fin)

        j=6
        i=6
        a=0
        b=0
        total=0
        total_e=0
        for di in detalle_ingreso:
            i = i + 1
            a=a+1

            h.cell(row=i, column=2).value = a
            h.cell(row=i, column=3).value = di.fechahora
            h.cell(row=i, column=4).value = di.get_tipo_display()
            h.cell(row=i, column=5).value = di.target
            h.cell(row=i, column=6).value = di.monto
            h.cell(row=i, column=7).value = di.descripcion
            total = total + di.monto



            ### BORDER CADA FILA ###
            h.cell(row=i, column=2).border = Border(top=Side(border_style='thin', color='FF000000'),
                                                    right=Side(border_style='thin', color='FF000000'),
                                                    bottom=Side(border_style='thin', color='FF000000'),
                                                    left=Side(border_style='thin', color='FF000000'))

            h.cell(row=i, column=3).border = Border(top=Side(border_style='thin', color='FF000000'),
                                                    right=Side(border_style='thin', color='FF000000'),
                                                    bottom=Side(border_style='thin', color='FF000000'),
                                                    left=Side(border_style='thin', color='FF000000'))

            h.cell(row=i, column=4).border = Border(top=Side(border_style='thin', color='FF000000'),
                                                    right=Side(border_style='thin', color='FF000000'),
                                                    bottom=Side(border_style='thin', color='FF000000'),
                                                    left=Side(border_style='thin', color='FF000000'))

            h.cell(row=i, column=5).border = Border(top=Side(border_style='thin', color='FF000000'),
                                                    right=Side(border_style='thin', color='FF000000'),
                                                    bottom=Side(border_style='thin', color='FF000000'),
                                                    left=Side(border_style='thin', color='FF000000'))

            h.cell(row=i, column=6).border = Border(top=Side(border_style='thin', color='FF000000'),
                                                    right=Side(border_style='thin', color='FF000000'),
                                                    bottom=Side(border_style='thin', color='FF000000'),
                                                    left=Side(border_style='thin', color='FF000000'))

            h.cell(row=i, column=7).border = Border(top=Side(border_style='thin', color='FF000000'),
                                                    right=Side(border_style='thin', color='FF000000'),
                                                    bottom=Side(border_style='thin', color='FF000000'),
                                                    left=Side(border_style='thin', color='FF000000'))


        i = i + 2
        h.cell(row=i, column=5).value = "TOTAL"
        h.cell(row=i, column=6).value = total


        for de in detalle_egreso:
            j = j + 1
            b=b+1

            h.cell(row=j, column=9).value = b
            h.cell(row=j, column=10).value = de.fechahora
            h.cell(row=j, column=11).value = de.get_tipo_display()
            h.cell(row=j, column=12).value = de.target
            h.cell(row=j, column=13).value = de.monto
            h.cell(row=j, column=14).value = de.descripcion
            total_e = total_e + de.monto



            ### BORDER CADA FILA ###
            h.cell(row=j, column=9).border = Border(top=Side(border_style='thin', color='FF000000'),
                                                    right=Side(border_style='thin', color='FF000000'),
                                                    bottom=Side(border_style='thin', color='FF000000'),
                                                    left=Side(border_style='thin', color='FF000000'))

            h.cell(row=j, column=10).border = Border(top=Side(border_style='thin', color='FF000000'),
                                                    right=Side(border_style='thin', color='FF000000'),
                                                    bottom=Side(border_style='thin', color='FF000000'),
                                                    left=Side(border_style='thin', color='FF000000'))

            h.cell(row=j, column=11).border = Border(top=Side(border_style='thin', color='FF000000'),
                                                    right=Side(border_style='thin', color='FF000000'),
                                                    bottom=Side(border_style='thin', color='FF000000'),
                                                    left=Side(border_style='thin', color='FF000000'))

            h.cell(row=j, column=12).border = Border(top=Side(border_style='thin', color='FF000000'),
                                                    right=Side(border_style='thin', color='FF000000'),
                                                    bottom=Side(border_style='thin', color='FF000000'),
                                                    left=Side(border_style='thin', color='FF000000'))

            h.cell(row=j, column=13).border = Border(top=Side(border_style='thin', color='FF000000'),
                                                    right=Side(border_style='thin', color='FF000000'),
                                                    bottom=Side(border_style='thin', color='FF000000'),
                                                    left=Side(border_style='thin', color='FF000000'))

            h.cell(row=j, column=14).border = Border(top=Side(border_style='thin', color='FF000000'),
                                                    right=Side(border_style='thin', color='FF000000'),
                                                    bottom=Side(border_style='thin', color='FF000000'),
                                                    left=Side(border_style='thin', color='FF000000'))


        j = j + 2
        h.cell(row=j, column=9).value = "TOTAL"
        h.cell(row=j, column=10).value = total_e



        response = HttpResponse(content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename=finanzas_vs.xls'
        libro.save(response)
        return response






    else:
        print(request.POST['tipo_movimiento'])
        libro = Workbook()
        libro = load_workbook("./detallejornada.xlsx")
        h = libro.get_sheet_by_name("Hoja1")

        j=Jornada.objects.get(id=id)


        h.cell(row=4, column=3).value = j.caja.descripcion
        h.cell(row=7, column=3).value = j.monto_actual
        h.cell(row=4, column=5).value = j.fechahora_inicio
        h.cell(row=4, column=7).value = j.fechahora_fin
        h.cell(row=7, column=5).value = j.asignado_inicio.username
        if(j.asignado_fin_id == None):
            h.cell(row=7, column=7).value = ""
        else:
            h.cell(row=7, column=7).value = j.asignado_fin.username

        fecha_inicio = datetime.strptime(request.POST['fechahora_inicio1'], '%d/%m/%Y %H:%M')
        fecha_fin = datetime.strptime(request.POST['fechahora_inicio2'], '%d/%m/%Y %H:%M')

        detalle=DetalleJornada.objects.filter(jornada_id=id,tipo=request.POST['tipo_movimiento'],fechahora__gte=fecha_inicio, fechahora__lte=fecha_fin)
        i=11
        a=0
        total=0
        for d in detalle:
            i = i + 1
            a=a+1

            h.cell(row=i, column=2).value = a
            h.cell(row=i, column=3).value = d.fechahora
            h.cell(row=i, column=4).value = d.get_tipo_display()
            h.cell(row=i, column=5).value = d.target
            h.cell(row=i, column=6).value = d.monto
            h.cell(row=i, column=7).value = d.descripcion
            total = total + d.monto



            ### BORDER CADA FILA ###
            h.cell(row=i, column=2).border = Border(top=Side(border_style='thin', color='FF000000'),
                                                    right=Side(border_style='thin', color='FF000000'),
                                                    bottom=Side(border_style='thin', color='FF000000'),
                                                    left=Side(border_style='thin', color='FF000000'))

            h.cell(row=i, column=3).border = Border(top=Side(border_style='thin', color='FF000000'),
                                                    right=Side(border_style='thin', color='FF000000'),
                                                    bottom=Side(border_style='thin', color='FF000000'),
                                                    left=Side(border_style='thin', color='FF000000'))

            h.cell(row=i, column=4).border = Border(top=Side(border_style='thin', color='FF000000'),
                                                    right=Side(border_style='thin', color='FF000000'),
                                                    bottom=Side(border_style='thin', color='FF000000'),
                                                    left=Side(border_style='thin', color='FF000000'))

            h.cell(row=i, column=5).border = Border(top=Side(border_style='thin', color='FF000000'),
                                                    right=Side(border_style='thin', color='FF000000'),
                                                    bottom=Side(border_style='thin', color='FF000000'),
                                                    left=Side(border_style='thin', color='FF000000'))

            h.cell(row=i, column=6).border = Border(top=Side(border_style='thin', color='FF000000'),
                                                    right=Side(border_style='thin', color='FF000000'),
                                                    bottom=Side(border_style='thin', color='FF000000'),
                                                    left=Side(border_style='thin', color='FF000000'))

            h.cell(row=i, column=7).border = Border(top=Side(border_style='thin', color='FF000000'),
                                                    right=Side(border_style='thin', color='FF000000'),
                                                    bottom=Side(border_style='thin', color='FF000000'),
                                                    left=Side(border_style='thin', color='FF000000'))


        i = i + 2
        h.cell(row=i, column=5).value = "TOTAL"
        h.cell(row=i, column=6).value = total


        response = HttpResponse(content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename=detallejornada'+str(j.id)+'.xls'
        libro.save(response)
        return response